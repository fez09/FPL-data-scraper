# FPL DATA FETCHER - May 2019
# THIS IS FOR THE 2019 - 2020 Season
# This code extracts the score history from Fantasy Premier League website
# Using the json data
# and exports it to an Excel Workbook
# https://www.fezfiles.com/fpl-data-fetcher
# https://github.com/fez09/FPL-data-scraper
#

from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import requests
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import IconSet, Rule, FormatObject
import base64
from urllib.request import urlopen
from os import path


## Create class

class fantasypl():

    ## Initialize GUI
    def __init__(self, parent):

        self.Frame = ttk.Frame(parent)
        self.Frame.grid()

        parent.resizable(False, False)
        parent.title('FPL Data Fetcher - v3')

        image_url = "http://i.imgur.com/QoNiPLP.gif"
        image_byt = urlopen(image_url).read()
        image_b64 = base64.encodebytes(image_byt)

        ## Creating and placing widgets
        self.logo = PhotoImage(data=image_b64)
        ttk.Label(self.Frame, image=self.logo).grid(row=1)
        ttk.Label(self.Frame, text="Enter your FPL ID").grid(row=2, padx=5, pady=5)
        ttk.Label(self.Frame, text="Data is for 2019/2020").grid(row=5, padx=5, pady=5)
        self.fpl_prompt = ttk.Entry(self.Frame, width=25, font=('Times New Roman', 16), justify=CENTER)
        self.fpl_prompt.grid(row=3, padx=5, pady=5)

        button1 = ttk.Button(self.Frame, text="Submit", command=self.button_press)
        button1.grid(row=4, padx=5, pady=5)
        self.style = ttk.Style()
        self.style.configure('TFrame', background='#00FF7C')
        self.style.configure('TButton', background='#00FF7C', font=('Calibri', 16))
        self.style.configure('TLabel', background='#00FF7C', font=('Calibri', 15, 'bold'))

    ## Combined function for button press
    def button_press(self):

        ## Check to see if a valid FPL ID is entered (Numbers only)
        if len(self.fpl_prompt.get()) != 0:
            if self.fpl_prompt.get().isdigit():
                if int(self.fpl_prompt.get()) < 6324237:  # Total number of FPL Players. Must know this value.
                    self.popup()
                    self.submit()
                else:
                    messagebox.showinfo(title="Error", message="Please enter a valid FPL ID")
                    self.fpl_prompt.delete(0, 'end')
                    return
            else:
                messagebox.showinfo(title="Error", message="Please enter a valid FPL ID")
                self.fpl_prompt.delete(0, 'end')
                return
        else:
            messagebox.showinfo(title="Error", message="Please enter a valid FPL ID")
            self.fpl_prompt.delete(0, 'end')
            return

    ## Loading Label after clicking submit
    def popup(self):
        label2 = ttk.Label(self.Frame, text="Please be patient...")
        label2.grid(row=2, padx=5, pady=5)
        label3 = ttk.Label(self.Frame, text="App might not respond")
        label3.grid(row=5, padx=5, pady=5)
        self.Frame.update_idletasks()
        label2.grid_forget()
        label3.grid_forget()

    ## Actions after clicking submit (importing data and creating workbook)
    def submit(self):

        ## Create workbook and sheets
        global g_w
        wb = openpyxl.Workbook()
        sheet0 = wb.create_sheet(index=0, title='Read_Me')
        sheet1 = wb.create_sheet(index=1, title='2019_2020')

        ## Create Read me sheet
        sheet0['B2'].value = 'Hey all. This excel file is the result of a python script'
        sheet0['B4'].value = 'The script uses the FPL API and json data to import your history from the ' \
                             'website and then exports it to this file '
        sheet0[
            'B6'].value = 'See more info at my site https://www.fezfiles.com/fpl-data-fetcher. Report bugs, contact at ten.dimensions10@gmail.com or Twitter: https://twitter.com/fez9o'
        sheet0['B8'].value = 'be aware that the code is raw and a lof of improvements can be made.'
        sheet0['B10'].value = 'Your data is in the next sheet. Change sheet tabs below or hold "CTRL+PgDown"'

        ## Import history JSON data
        url1 = 'https://fantasy.premierleague.com/api/entry/{}/history/'.format(self.fpl_prompt.get())
        url2 = 'https://fantasy.premierleague.com/api/bootstrap-static/'
        json_history = requests.get(url1).json()
        json_live = requests.get(url2).json()
        num_of_gw = len(json_history['current'])
        participants = json_live['total_players']

        ## Import league data and team name from new url due to change in FPL api since 2019-2020 Season
        url9 = 'https://fantasy.premierleague.com/api/entry/{}/'.format(self.fpl_prompt.get())
        json_info = requests.get(url9).json()
        team_name = json_info['name']

        ## Import gameweek history and insert data in sheet
        header1 = ['GW', 'GP', 'GW AVG', 'GW HS', 'PB', 'TM', 'TC', 'GR', 'PGR', 'OP', 'OR', 'POR', 'Position', 'TV']
        headerrow = 1
        for key in range(14):
            sheet1.cell(row=headerrow, column=key + 3).value = str(header1[key])

        o_r1 = []  # To make a list of overall rank for inserting change in rank symbols
        for each in json_history['current']:
            g_w = each['event']
            points = each['points']
            p_b = each['points_on_bench']
            t_m = each['event_transfers']
            t_c = each['event_transfers_cost']
            g_w_r = each['rank']
            o_r = each['overall_rank']
            t_v = each['value']
            o_r1.append(o_r)    # This is for creating rank symbols in the excel sheet
            o_p = each['total_points']

            p_g_r = 100 - (((participants - g_w_r) / participants) * 100)
            p_o_r = 100 - (((participants - o_r) / participants) * 100)
            p_o_s = 1  # placeholder which will be replaced later down the code

            history_list = [g_w, points, p_b, t_m, t_c, g_w_r, p_g_r, o_p, o_r, p_o_r, p_o_s, t_v / 10]
            for rownum in range(g_w + 1, g_w + 2):
                sheet1.cell(row=rownum, column=3).value = g_w
            for rownum in range(g_w + 1, g_w + 2):
                sheet1.cell(row=rownum, column=4).value = points
            for rownum in range(g_w + 1, g_w + 2):
                for key in range(2, 12):
                    sheet1.cell(row=rownum, column=key + 5).value = history_list[key]

        ## Appending overall rank changes to a list so it becomes easier to make symbols in the table
        p_o_s_1 = []
        p_o_s_1.append(0)
        for count in range(1, g_w):
            if o_r1[count - 1] > o_r1[count]:
                p_o_s = 1
            elif o_r1[count - 1] == o_r1[count]:
                p_o_s = 0
            else:
                p_o_s = -1
            p_o_s_1.append(p_o_s)

        for each in json_history['current']:
            g_w = each['event']
            sheet1.cell(row=g_w + 1, column=15).value = p_o_s_1[g_w - 1]

        ## Import gameweek average points and highest points
        for each in json_live['events']:
            g_w = each['id']
            h_p = each['highest_score']
            a_v_g = each['average_entry_score']
            for rownum in range(g_w + 1, g_w + 2):
                sheet1.cell(row=rownum, column=5).value = a_v_g
            for rownum in range(g_w + 1, g_w + 2):
                sheet1.cell(row=rownum, column=6).value = h_p

        ## Import all football players' data in premier league
        player_d = {}
        for each in json_live['elements']:
            pl_position = each['element_type']
            pl_id = each['id']
            pl_name = each['web_name']
            player_d[pl_id] = pl_name

        tot_player = len(player_d)

        ## Select data for Chip usage and enter in GW history as highlights
        wildcardfill = PatternFill(start_color='ffff0000', end_color='ffff0000', fill_type='solid')
        freehitfill = PatternFill(start_color='ffff00ff', end_color='ffff00ff', fill_type='solid')
        bboostfill = PatternFill(start_color='ffffa500', end_color='ffffa500', fill_type='solid')
        triplecapfill = PatternFill(start_color='ff0099ff', end_color='ff0099ff', fill_type='solid')
        gwh_col = range(3, 17)
        for each in json_history['chips']:
            chipgw = each['event']
            chip = each['name']
            while chip == 'wildcard':
                for key in gwh_col:
                    wc = sheet1.cell(row=chipgw + 1, column=key)
                    wc.fill = wildcardfill
                break
            while chip == 'bboost':
                for key in gwh_col:
                    wc = sheet1.cell(row=chipgw + 1, column=key)
                    wc.fill = bboostfill
                break
            while chip == 'freehit':
                for key in gwh_col:
                    wc = sheet1.cell(row=chipgw + 1, column=key)
                    wc.fill = freehitfill
                break
            while chip == '3xc':
                for key in gwh_col:
                    wc = sheet1.cell(row=chipgw + 1, column=key)
                    wc.fill = triplecapfill
                break

        ## Import weekly team player data and points
        gwteamheaderow = 50 #41+9
        gwtitle = 3
        for gw in range(1, 48): #39+9
            sheet1.cell(row=gwteamheaderow, column=gwtitle).value = str('GW {}'.format(gw))
            sheet1.cell(row=gwteamheaderow, column=gwtitle + 1).value = str('P {}'.format(gw))
            gwtitle = gwtitle + 2
        capfill = PatternFill(start_color='ff15dd43', end_color='ff15dd43', fill_type='solid')
        vcapfill = PatternFill(start_color='ff00FFDA', end_color='ff00FFDA', fill_type='solid')
        benchfill = PatternFill(start_color='ffBA6B12', end_color='ffBA6B12', fill_type='solid')
        for rownum in range(62, 66): # 62 to 66
            for colnum in range(3, 97): # 79+18 to 97
                bench = sheet1.cell(row=rownum, column=colnum)
                bench.fill = benchfill

        gwteamcol = ((39+9) * 2 - int(num_of_gw * 2)) - 1  # To accomodate people who started late. Don't ask how I came up with this formula.

        capfont = Font(underline='single')
        # gwteamcol = 1
        for each in json_history['current']:
            g_w = each['event']
            url3 = 'https://fantasy.premierleague.com/api/entry/{}/event/{}/picks/'.format(self.fpl_prompt.get(), g_w)
            json_pick = requests.get(url3).json()
            gwteamcol = gwteamcol + 2
            gwteamrow = 42+9
            url4 = 'https://fantasy.premierleague.com/api/event/{}/live/'.format(g_w)
            json_points = requests.get(url4).json()
            total_players = len(json_points['elements'])

            for each1 in json_pick['picks']:
                player_id = each1['element']
                player_idnew = int(player_id)
                captain = each1['is_captain']
                vicecapt = each1['is_vice_captain']
                multiplier = each1['multiplier']
                pl_name = player_d[player_id]
                plist = {player_id: pl_name}

                # Assigning points to selected players
                for each2 in json_points['elements']:
                    pointsvid = each2['id']
                    for t1 in range(0, total_players):
                        if player_idnew == json_points['elements'][t1]['id']:
                            pl_points = json_points['elements'][t1]['stats']['total_points']
                            sheet1.cell(row=gwteamrow, column=gwteamcol + 1).value = pl_points
                            if multiplier == 2:
                                sheet1.cell(row=gwteamrow, column=gwteamcol + 1).value = pl_points * 2
                            elif multiplier == 3:
                                sheet1.cell(row=gwteamrow, column=gwteamcol + 1).value = pl_points * 3
                            break

                # Assigning captaincy and vice-captaincy
                for values in plist.values():
                    sheet1.cell(row=gwteamrow, column=gwteamcol).value = values
                if captain == True:
                    capf = sheet1.cell(row=gwteamrow, column=gwteamcol)
                    capf.fill = capfill
                    capf.font = capfont
                if vicecapt == True:
                    vcapf = sheet1.cell(row=gwteamrow, column=gwteamcol)
                    vcapf.fill = vcapfill
                gwteamrow = gwteamrow + 1


        ## Import classic league history
        sheet1.merge_cells('BR1:BS1')  # League Rank header
        sheet1['BR1'].value = 'League Rank History'

        clrow = 2
        num_of_leagues = len(json_info['leagues']['classic'])
        clheader = ['League Name', 'Rank']
        for leaguecolumn in range(2):
            sheet1.cell(row=clrow, column=leaguecolumn + 70).value = str(clheader[leaguecolumn])
        for each in json_info['leagues']['classic']:
            leaguename = each['name']
            leagueposition = each['entry_rank']
            league_data = [leaguename, leagueposition]
            clrow = clrow + 1
            for clcol in range(2):
                sheet1.cell(row=clrow, column=clcol + 70).value = league_data[clcol]

        ## Import Cup History
        sheet1.merge_cells('BV1:BZ1')  # Cup History Header
        sheet1['BV1'].value = 'FPL Cup History'

        url8 = 'https://fantasy.premierleague.com/api/entry/{}/cup/'.format(self.fpl_prompt.get())
        json_cup = requests.get(url8).json()
        num_of_cups = len(json_cup['cup_matches'])
        cuplist = ['GW', 'Team Name 1', 'Points 1', 'Team Name 2', 'Points 2 ']
        cuprow = 2
        for col in range(5):
            sheet1.cell(row=cuprow, column=col + 74).value = str(cuplist[col])

        if num_of_cups > 0:
            for each in json_cup['cup_matches']:
                cupgw = each['event']
                entry_1 = each['entry_1_name']
                entry_2 = each['entry_2_name']
                entrypoints_1 = each['entry_1_points']
                entrypoints_2 = each['entry_2_points']
                cup_data = [cupgw, entry_1, entrypoints_1, entry_2, entrypoints_2]
                cuprow = cuprow + 1
                for colnum in range(5):
                    sheet1.cell(row=cuprow, column=colnum + 74).value = cup_data[colnum]
        else:
            sheet1.cell(row=3, column=74).value = "Failed to qualify for the cup"

        ## Import h2h details
        sheet1.merge_cells('BN1:BO1')  # H2H Team Header
        sheet1['BN1'].value = 'H2H History'
        num_of_h2h = len(json_info['leagues']['h2h'])

        h2h_header = ['H2H League', 'Rank']
        h2hrow = 2
        for h2hcol in range(2):
            sheet1.cell(row=h2hrow, column=h2hcol + 66).value = str(h2h_header[h2hcol])

        if num_of_h2h > 0:
            for each in json_history['leagues']['h2h']:
                h2hname = each['name']
                h2hrank = each['entry_rank']
                h2h_data = [h2hname, h2hrank]
                h2hrow = h2hrow + 1
                for colnum in range(2):
                    sheet1.cell(row=h2hrow, column=colnum + 66).value = h2h_data[colnum]
        else:
            sheet1.cell(row=3, column=66).value = "No H2H leagues entered." \
 \
        ## Import Gameweek Transfer history
        sheet1.merge_cells('CV1:CZ1')
        sheet1['CV1'].value = 'Transfer History'

        transferheader = ['GW', 'Transfer In', 'Value In ', 'Transfer Out', 'Value Out']
        transferhrow = 2
        for tkey in range(5):
            sheet1.cell(row=transferhrow, column=tkey + 100).value = str(transferheader[tkey]) #82+18
        url5 = 'https://fantasy.premierleague.com/api/entry/{}/transfers/'.format(self.fpl_prompt.get())
        json_transfer = requests.get(url5).json()
        gwtransferrow = 2
        gwtransfercol = 82+18
        num_of_t = len(json_transfer)

        if num_of_t == 0:
            sheet1.cell(row=gwtransferrow + 1, column=gwtransfercol).value = "No Transfers Made"
        else:
            for each in json_transfer:
                transferin = each['element_in']
                transferout = each['element_out']
                incost = each['element_in_cost']
                outcost = each['element_out_cost']
                transfergw = each['event']
                t_in_name = player_d.get(transferin, 0)
                t_out_name = player_d.get(transferout, 0)
                trans_data = [transfergw, t_in_name, incost / 10, t_out_name, outcost / 10]
                gwtransferrow = gwtransferrow + 1
                for colnum in range(5):
                    sheet1.cell(row=gwtransferrow, column=colnum + gwtransfercol).value = trans_data[colnum]

        ## Import Overall Dream Team Data
        sheet1.merge_cells('BI1:BJ1')  # Dream Team Header
        sheet1['BI1'].value = str('Overall Dream Team')

        overalldtheader = ['Player Name', 'Total Points']
        for odtcol in range(2):
            sheet1.cell(row=2, column=odtcol + 61).value = str(overalldtheader[odtcol])

        url6 = 'https://fantasy.premierleague.com/api/dream-team/'
        json_dreamteam = requests.get(url6).json()
        dtrow = 2

        for each in json_dreamteam['team']:
            dtpoints = each['points']
            dtplayer = each['element']
            dt_name = player_d.get(dtplayer, 0)
            dt_data = [dt_name, dtpoints]
            dtrow = dtrow + 1
            for colnum in range(2):
                sheet1.cell(row=dtrow, column=colnum + 61).value = dt_data[colnum]

        ## Import Weekly Dream Team Data
        dtteamheaderrow = 59+9
        dttitle = 3
        for dt in range(1, 39+9):
            sheet1.cell(row=dtteamheaderrow, column=dttitle).value = str('GW {}'.format(dt))
            sheet1.cell(row=dtteamheaderrow, column=dttitle + 1).value = str('P {}'.format(dt))
            dttitle = dttitle + 2

        dtteamcol = 1
        for each in range(1, 39+9):  ## CHANGE THIS BACK TO 39 TO ACOMMODATE ALL GAMEWEEKS
            url7 = 'https://fantasy.premierleague.com/api/dream-team/{}/'.format(each)
            json_weeklydt = requests.get(url7).json()
            # print(json_weeklydt)
            dtteamrow = 69 #60+9
            dtteamcol = dtteamcol + 2
            for each1 in json_weeklydt['team']:
                dtpl_id = each1['element']
                dt_points = each1['points']
                pl_name = player_d[dtpl_id]
                dtlist = {pl_name: dt_points}
                for values in dtlist.values():
                    sheet1.cell(row=dtteamrow, column=dtteamcol + 1).value = values
                for values2 in dtlist.keys():
                    sheet1.cell(row=dtteamrow, column=dtteamcol).value = values2
                dtteamrow = dtteamrow + 1

        ## Creating Legend
        legendlist = ['Legend', 'Wildcard', 'Benchboost', 'Triple Captain', 'Free Hit', 'Captain', 'Vice Captain',
                      'Bench']
        legendrow = 5
        for lkey in range(8):
            sheet1.cell(row=lkey + legendrow, column=1).value = legendlist[lkey]
            lewc = sheet1.cell(row=legendrow + 1, column=1)
            lewc.fill = wildcardfill
            lebb = sheet1.cell(row=legendrow + 2, column=1)
            lebb.fill = bboostfill
            letc = sheet1.cell(row=legendrow + 3, column=1)
            letc.fill = triplecapfill
            lefh = sheet1.cell(row=legendrow + 4, column=1)
            lefh.fill = freehitfill
            leca = sheet1.cell(row=legendrow + 5, column=1)
            leca.fill = capfill
            levca = sheet1.cell(row=legendrow + 6, column=1)
            levca.fill = vcapfill
            leben = sheet1.cell(row=legendrow + 7, column=1)
            leben.fill = benchfill

        ## Creating Team name and FPL ID
        sheet1['A2'].value = 'FPL ID: {}'.format(self.fpl_prompt.get())
        sheet1['A1'].value = 'Team: {}'.format(team_name)
        sheet1['A3'].value = 'Players: {}'.format(participants)

        ## Cell Styling
        headerfont = Font(bold=True)
        alignment = Alignment(horizontal='center')

        for key in range(74, 79):  # FPL CUP 'GW/Team Name/Points/Team Name/Points'
            row2 = sheet1.cell(row=2, column=key)
            row2.font = headerfont
            row2.alignment = alignment

        for row in range(1, num_of_cups + 3):  # FPL CUP History details
            for col in range(74, 79):
                cup1 = sheet1.cell(row=row, column=col)
                cup1.alignment = alignment

        cup2 = sheet1.cell(row=1, column=74)  # 'FPL CUP History'
        cup2.font = headerfont
        cup2.alignment = alignment

        for key in range(3, 17):  # 'GW/GP/GW AVG/GW HS/PB/......'
            row1 = sheet1.cell(row=1, column=key)
            row1.font = headerfont
            row1.alignment = alignment

        for key in range(3, 97):  # GW Teams 'GW1/P1/GW2/P2.....' 79+18, 41+9
            row41 = sheet1.cell(row=50, column=key)
            row41.font = headerfont
            row41.alignment = alignment

        for key1 in range(2, 49):  # GW history full table 40+9
            for key2 in range(3, 10):
                set1 = sheet1.cell(row=key1, column=key2)
                set1.alignment = alignment

        for key1 in range(51, 66):  # Team history full table 42+9, 57+9, 79+18
            for key2 in range(3, 97):
                set2 = sheet1.cell(row=key1, column=key2)
                set2.alignment = alignment

        for key in range(2, 49):  # GW history table value format 40+9
            col1 = sheet1.cell(row=key, column=11)  # Percentile GW Rank
            col1.number_format = '0.00000'
            col1.alignment = alignment
            col2 = sheet1.cell(row=key, column=14)  # Percentile Overall Rank
            col2.number_format = '0.00000'
            col2.alignment = alignment
            col3 = sheet1.cell(row=key, column=16)  # Team Value
            col3.number_format = '0.0'
            col3.alignment = alignment
            col4 = sheet1.cell(row=key, column=12)  # Overall Points
            col4.alignment = alignment
            col5 = sheet1.cell(row=key, column=10)  # Gameweek Rank
            col5.number_format = '#,##0'
            col5.alignment = alignment
            col6 = sheet1.cell(row=key, column=13)  # Overall Rank
            col6.number_format = '#,##0'
            col6.alignment = alignment
            col4 = sheet1.cell(row=key, column=15)  # Position
            col4.alignment = alignment

        for key1 in range(1, 3 + num_of_t):  # Transfer history table 82+18, 88+18
            for key2 in range(100, 106):
                set3 = sheet1.cell(row=key1, column=key2)
                set3.alignment = alignment
        for key in range(100, 106):
            set4 = sheet1.cell(row=1, column=key)
            set4.alignment = alignment
            set4.font = headerfont

        for key in range(100, 106):  # Transfer history 'GW/Transfer in/Value....'
            row1 = sheet1.cell(row=1, column=key)
            row1.font = headerfont

        for key1 in range(2, num_of_t + 1):  # Transfer table value In 84+18
            col7 = sheet1.cell(row=key1, column=102)
            col7.number_format = '0.0'

        for key1 in range(2, num_of_t + 1):  # Transfer table value Out 86+18
            col7 = sheet1.cell(row=key1, column=104)
            col7.number_format = '0.0'

        for col8 in range(70, 73):  # League Rank table
            lrh = sheet1.cell(row=1, column=col8)
            lrh.font = headerfont
            for row8 in range(2, num_of_leagues + 3):
                lr = sheet1.cell(row=row8, column=col8)
                lr.alignment = alignment

        lr1 = sheet1.cell(row=1, column=70)  # League Rank title
        lr1.font = headerfont
        lr1.alignment = alignment

        for col9 in range(3, 97):  # Gameweek Dream Team full table and title 79+18, 59+9
            gwdt = sheet1.cell(row=68, column=col9)
            gwdt.font = headerfont
            for row9 in range(68, 80): # 59+9, 71+9
                gwdtt = sheet1.cell(row=row9, column=col9)
                gwdtt.alignment = alignment

        for col10 in range(61, 63):  # Overall dream team table and titles
            odt = sheet1.cell(row=1, column=col10)
            odt.font = headerfont
            odt.alignment = alignment
        for col10 in range(61, 63):
            for row10 in range(2, 14):
                odtt = sheet1.cell(row=row10, column=col10)
                odtt.alignment = alignment
        for col10 in range(61, 63):
            odt = sheet1.cell(row=2, column=col10)
            odt.font = headerfont

        for col11 in range(66, 68):  # Head2Head table and titles
            h2ht = sheet1.cell(row=1, column=col11)
            h2ht.font = headerfont
            h2ht.alignment = alignment
        for col11 in range(66, 68):
            for row11 in range(2, num_of_h2h + 3):
                h2htt = sheet1.cell(row=row11, column=col11)
                h2htt.alignment = alignment
        for col11 in range(66, 68):
            h2ht = sheet1.cell(row=2, column=col11)
            h2ht.font = headerfont

        ## Creating Position symbols for GW history p_o_s using p_o_s_1
        first = FormatObject(type='num', val=-1)
        second = FormatObject(type='num', val=0)
        third = FormatObject(type='num', val=1)
        iconset = IconSet(iconSet='3Arrows', cfvo=[first, second, third], showValue=None, percent=None,
                          reverse=None)
        rule = Rule(type='iconSet', iconSet=iconset)
        sheet1.conditional_formatting.add('O2:O48', cfRule=rule) #39+9

        ## Creating Tables
        table1 = Table(displayName='GWH', ref='C1:P48')  # GW History 39+9
        style1 = TableStyleInfo(name="TableStyleMedium11", showRowStripes=True)
        table1.tableStyleInfo = style1
        sheet1.add_table(table1)

        table2 = Table(displayName='GWT', ref='C50:CR65')  # Team History 41+9, BZ to CR 56+ 9
        style2 = TableStyleInfo(name="TableStyleLight15", showRowStripes=True)
        table2.tableStyleInfo = style2
        sheet1.add_table(table2)

        if num_of_t > 0:
            num_of_trow = num_of_t
        else:
            num_of_trow = 1
        table3 = Table(displayName='TH', ref='CV2:CZ{}'.format(num_of_trow + 2))  # Transfer History
        style3 = TableStyleInfo(name="TableStyleMedium12", showRowStripes=True)
        table3.tableStyleInfo = style3
        sheet1.add_table(table3)

        if num_of_cups > 0:
            cup_table_row = 2
        else:
            cup_table_row = 3

        table4 = Table(displayName='CH', ref='BV2:BZ{}'.format(num_of_cups + cup_table_row))  # FPL Cup history
        style4 = TableStyleInfo(name="TableStyleMedium13", showRowStripes=True)
        table4.tableStyleInfo = style4
        sheet1.add_table(table4)

        table5 = Table(displayName='CLR', ref='BR2:BS{}'.format(num_of_leagues + 2))  # Classic League Rank
        style5 = TableStyleInfo(name="TableStyleMedium10", showRowStripes=True)
        table5.tableStyleInfo = style5
        sheet1.add_table(table5)

        if num_of_h2h > 0:
            h2h_table_row = 2
        else:
            h2h_table_row = 3

        table6 = Table(displayName='HTOH', ref='BN2:BO{}'.format(num_of_h2h + h2h_table_row))  # H2H Rank
        style6 = TableStyleInfo(name="TableStyleMedium11", showRowStripes=True)
        table6.tableStyleInfo = style6
        sheet1.add_table(table6)

        table7 = Table(displayName='ODT', ref='BI2:BJ13')  # Overall Dream Team
        style7 = TableStyleInfo(name="TableStyleMedium7", showRowStripes=True)
        table7.tableStyleInfo = style7
        sheet1.add_table(table7)

        table8 = Table(displayName='GWDT', ref='C68:CR79')  # GW Dream Team 59+9, 70+9
        style8 = TableStyleInfo(name="TableStyleLight17", showRowStripes=True)
        table8.tableStyleInfo = style8
        sheet1.add_table(table8)

        ## Creating Chart
        chart1 = LineChart()
        chart1.title = 'Gameweek Points / Average Points / Points Benched / Highest GW Score'
        data1 = Reference(sheet1, min_col=4, max_col=7, min_row=1, max_row=48) # 39+9
        chart1.height = 20
        chart1.width = 60
        chart1.add_data(data1, titles_from_data=True)
        sheet1.add_chart(chart1, "T2")

        ## Save workbook
        wb.save("FPL.Data.19-20.{}.xlsx".format(self.fpl_prompt.get()))

        ## Clear text box and show success dialog
        item_path = str(path.realpath("FPL.Data.19-20.{}.xlxs".format(self.fpl_prompt.get())))
        messagebox.showinfo(title="Success", message="Data for \'{}\' imported successfully.\n"
                                                     "File saved in {}".format(team_name, item_path))
        self.clear()

    ## Define function to clear dialog
    def clear(self):
        self.fpl_prompt.delete(0, 'end')


## Call GUI
def main():
    root = Tk()
    fantasypl(root)
    root.mainloop()


if __name__ == "__main__":
    main()
