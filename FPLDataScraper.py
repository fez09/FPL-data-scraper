# FPL ARCHIVE EXTRACTOR - May 2018
# THIS IS FOR THE 2017 - 2018 Season
# This code extracts the score history from Fantasy Premier League website
# Using the API and json data
# and exports it to an Excel Workbook
# https://github.com/fez09/FPL-data-scraper
# Compiled by Fez [u/CinnamonUranium]

from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import openpyxl
import requests
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import IconSet, Rule, FormatObject
import base64
from urllib.request import urlopen


# Create class
class fantasypl:

    # Initialize GUI
    def __init__(self, parent):

        self.Frame = ttk.Frame(parent)
        self.Frame.grid()

        parent.resizable(False, False)
        parent.title("FPL Data Scraper")

        image_url = "http://i.imgur.com/QoNiPLP.gif"
        image_byt = urlopen(image_url).read()
        image_b64 = base64.encodebytes(image_byt)

        # Creating and placing widgets
        self.logo = PhotoImage(data=image_b64)
        ttk.Label(self.Frame, image=self.logo).grid()
        ttk.Label(self.Frame, text="Enter your FPL ID").grid(row=1, padx=5, pady=5)
        ttk.Label(self.Frame, text="Please be patient while your data"
                                   " is imported").grid(row=5, padx=5, pady=5)
        self.fpl_prompt = ttk.Entry(self.Frame, width=25, font=('Times New Roman', 18), justify=CENTER)
        self.fpl_prompt.grid(row=3, padx=5, pady=5)
        ttk.Button(self.Frame, text="Submit", command=self.submit).grid(row=4, padx=5, pady=5)
        self.style = ttk.Style()
        self.style.configure('TFrame', background='#00FF7C')
        self.style.configure('TButton', background='#00FF7C', font=('Calibri', 15))
        self.style.configure('TLabel', background='#00FF7C', font=('Calibri', 15, 'bold'))

    # Actions after clicking submit
    def submit(self):

        print("Please wait while your data is being fetched. "
              "This may take some time depending on your computer/internet")

        # Import history JSON data
        url1 = 'https://fantasy.premierleague.com/drf/entry/{}/history'.format(self.fpl_prompt.get())
        url2 = 'https://fantasy.premierleague.com/drf/bootstrap-static'
        json_history = requests.get(url1).json()
        json_live = requests.get(url2).json()

        # Create workbook and sheets
        wb = openpyxl.Workbook()
        sheet0 = wb.create_sheet(index=0, title='Read_Me')
        sheet1 = wb.create_sheet(index=1, title='2017_2018')

        # Create Read me sheet
        sheet0['B2'].value = 'Hey all. This excel file is the result of a very simple and amateurish python script'
        sheet0['B4'].value = 'The script uses the FPL API and json data to import your history from the ' \
                             'website and then exports it to this file '
        sheet0['B6'].value = 'If you want to give me more motivation to improve the script, you can' \
                             ' buy me a drink at ten.dimensions10@gmail.com ' \
                             'or just donate to a charity of your choice'
        sheet0['B8'].value = 'If you are interested in the python code you can find it at ' \
                             'https://github.com/fez09/FPL-data-scraper'
        sheet0['B10'].value = 'be aware that the code is VERY amateurish and a lof of improvements can be made.'
        sheet0['B12'].value = 'Your data is in the next sheet. Change sheets below or hold "CTRL+PgDown"'

        # Create headers
        header1 = ['GW', 'GP', 'GW AVG', 'PB', 'TM', 'TC', 'GR', 'OP', 'OR', 'Position', 'TV']
        transferheader = ['GW', 'Transfer In', 'Value', 'Transfer Out', 'Value']
        sheet1.merge_cells('E58:F58')
        sheet1['E58'] = 'Overall Dream Team'
        for tkey in range(5):
            sheet1.cell(row=1, column=tkey + 42).value = transferheader[tkey]
        for key in range(11):
            sheet1.cell(row=1, column=key + 2).value = header1[key]
        for gw in range(1, 39):
            sheet1.cell(row=41, column=gw + 1).value = 'GW {}'.format(gw)

        # Import gameweek history and insert data in sheet
        for each in json_history['history']:
            g_w = each['event']
            points = each['points']
            p_b = each['points_on_bench']
            t_m = each['event_transfers']
            t_c = each['event_transfers_cost']
            g_w_r = each['rank']
            o_r = each['overall_rank']
            t_v = each['value']
            p_o_s = each['movement']
            o_p = each['total_points']
            if p_o_s == 'up':
                p_o_s = 1
            elif p_o_s == 'new':
                p_o_s = 0
            else:
                p_o_s = -1
            history_list = [g_w, points, p_b, t_m, t_c, g_w_r, o_p, o_r, p_o_s, t_v / 10]
            for rownum in range(g_w + 1, g_w + 2):
                sheet1.cell(row=rownum, column=2).value = g_w
            for rownum in range(g_w + 1, g_w + 2):
                sheet1.cell(row=rownum, column=3).value = points
            for rownum in range(g_w + 1, g_w + 2):
                for key in range(2, 10):
                    sheet1.cell(row=rownum, column=key + 3).value = history_list[key]

        # Import gameweek average points and insert data in sheet
        for each in json_live['events']:
            g_w = each['id']
            a_v_g = each['average_entry_score']
            for rownum in range(g_w + 1, g_w + 2):
                sheet1.cell(row=rownum, column=4).value = a_v_g

        # Import all player data in premier league
        d = {}
        for each in json_live['elements']:
            pl_id = each['id']
            pl_name = each['web_name']
            d[pl_id] = pl_name

        # Select team player data for personal team for each gameweek and enter in sheet
        colnum = 1
        for each in json_history['history']:
            g_w = each['event']
            url3 = 'https://fantasy.premierleague.com/drf/entry/{}/event/{}/picks'.format(self.fpl_prompt.get(), g_w)
            json_pick = requests.get(url3).json()
            colnum = colnum + 1
            rownum = 42
            # noinspection PyAssignmentToLoopOrWithParameter
            for each in json_pick['picks']:
                player_id = each['element']
                pl_name = d[player_id]
                plist = {player_id: pl_name}
                for values in plist.values():
                    sheet1.cell(row=rownum, column=colnum).value = values
                    rownum = rownum + 1
        startfill = PatternFill(start_color='ff15dd43', end_color='ff15dd43', fill_type='solid')
        benchfill = PatternFill(start_color='ff00FFDA', end_color='ff00FFDA', fill_type='solid')
        for rownum in range(42, 53):
            for colnum in range(2, 40):
                start = sheet1.cell(row=rownum, column=colnum)
                start.fill = startfill
        for rownum in range(53, 57):
            for colnum in range(2, 40):
                bench = sheet1.cell(row=rownum, column=colnum)
                bench.fill = benchfill

        # Import Gameweek Transfer data
        url4 = 'https://fantasy.premierleague.com/drf/entry/{}/transfers'.format(self.fpl_prompt.get())
        json_transfer = requests.get(url4).json()
        rownum = 1
        num_of_t = len(json_transfer['history'])
        for each in json_transfer['history']:
            transferin = each['element_in']
            transferout = each['element_out']
            incost = each['element_in_cost']
            outcost = each['element_out_cost']
            transfergw = each['event']
            t_in_name = d.get(transferin, 0)
            t_out_name = d.get(transferout, 0)
            trans_data = [transfergw, t_in_name, incost / 10, t_out_name, outcost / 10]
            rownum = rownum + 1
            for colnum in range(5):
                sheet1.cell(row=rownum, column=colnum + 42).value = trans_data[colnum]

        # Import Dream Team Data
        url5 = 'https://fantasy.premierleague.com/drf/dream-team'
        json_dreamteam = requests.get(url5).json()
        rownum2 = 58
        for each in json_dreamteam['team']:
            dtpoints = each['points']
            dtplayer = each['element']
            dt_name = d.get(dtplayer, 0)
            dt_data = [dt_name, dtpoints]
            rownum2 = rownum2 + 1
            for colnum in range(2):
                sheet1.cell(row=rownum2, column=colnum + 5).value = dt_data[colnum]

        # Select data for Chip usage and enter in excel as highlights
        wildcardfill = PatternFill(start_color='ffff0000', end_color='ffff0000', fill_type='solid')
        bboostfill = PatternFill(start_color='ffff00ff', end_color='ffff00ff', fill_type='solid')
        freehitfill = PatternFill(start_color='ffffa500', end_color='ffffa500', fill_type='solid')
        triplecapfill = PatternFill(start_color='ff0099ff', end_color='ff0099ff', fill_type='solid')
        for each in json_history['chips']:
            chipgw = each['event']
            chip = each['name']
            while chip == 'wildcard':
                for key in range(1, 12):
                    wc = sheet1.cell(row=chipgw + 1, column=key + 1)
                    wc.fill = wildcardfill
                break
            while chip == 'bboost':
                for key in range(1, 12):
                    wc = sheet1.cell(row=chipgw + 1, column=key + 1)
                    wc.fill = bboostfill
                break
            while chip == 'freehit':
                for key in range(1, 12):
                    wc = sheet1.cell(row=chipgw + 1, column=key + 1)
                    wc.fill = freehitfill
                break
            while chip == '3xc':
                for key in range(1, 12):
                    wc = sheet1.cell(row=chipgw + 1, column=key + 1)
                    wc.fill = triplecapfill
                break

        # Creating Legend
        legendlist = ['Legend', 'Wildcard', 'Benchboost', 'Triple Captain', 'Free Hit']
        for lkey in range(5):
            sheet1.cell(row=lkey + 59, column=2).value = legendlist[lkey]
            lewc = sheet1.cell(row=59 + 1, column=2)
            lewc.fill = wildcardfill
            lebb = sheet1.cell(row=59 + 2, column=2)
            lebb.fill = bboostfill
            letc = sheet1.cell(row=59 + 3, column=2)
            letc.fill = triplecapfill
            lefh = sheet1.cell(row=59 + 4, column=2)
            lefh.fill = freehitfill

        # Creating excel cell names
        alphabet = []
        for letter in range(65, 91):
            alphabet.append(chr(letter))
        alphabeta = []
        for letter in range(65, 91):
            alphabeta.append('A' + chr(letter))
        # AtoAZ = alphabet + alphabeta

        # Cell Styling
        headerfont = Font(bold=True)
        alignment = Alignment(horizontal='center')

        for key in range(1, 13):  # GW History Table header alignment
            row1 = sheet1.cell(row=1, column=key)
            row1.font = headerfont
            row1.alignment = alignment
        for key in range(1, 40):  # Team history table header alignment
            row41 = sheet1.cell(row=41, column=key)
            row41.font = headerfont
            row41.alignment = alignment
        for key1 in range(2, 40):  # GW history table alignment
            for key2 in range(2, 8):
                set1 = sheet1.cell(row=key1, column=key2)
                set1.alignment = alignment
        for key1 in range(42, 57):  # Team history table alignment
            for key2 in range(2, 40):
                set2 = sheet1.cell(row=key1, column=key2)
                set2.alignment = alignment
        for key in range(2, 40):  # GW history table alignment and value format
            col3 = sheet1.cell(row=key, column=12)
            col3.number_format = '0.0'
            col3.alignment = alignment
            col4 = sheet1.cell(row=key, column=9)
            col4.alignment = alignment
            col5 = sheet1.cell(row=key, column=8)
            col5.number_format = '#,##0'
            col6 = sheet1.cell(row=key, column=10)
            col6.number_format = '#,##0'
            col4 = sheet1.cell(row=key, column=11)
            col4.alignment = alignment
        for key1 in range(1, 3 + num_of_t):  # Transfer table alignment
            for key2 in range(42, 48):
                set3 = sheet1.cell(row=key1, column=key2)
                set3.alignment = alignment
        for key in range(42, 48):  # Transfer table header
            row1 = sheet1.cell(row=1, column=key)
            row1.font = headerfont
        for key1 in range(2, num_of_t + 1):  # Transfer table value format
            col7 = sheet1.cell(row=key1, column=45)
            col7.number_format = '0.0'
        for key1 in range(2, num_of_t + 1):  # Transfer table value format
            col7 = sheet1.cell(row=key1, column=47)
            col7.number_format = '0.0'
        for key1 in range(58, 70):  # Dream Team Table
            for key2 in range(5, 7):
                col8 = sheet1.cell(row=key1, column=key2)
                col8.alignment = alignment
        col9 = sheet1.cell(row=58, column=5)  # Dream Team Table header
        col9.font = headerfont
        dreamteamfill = PatternFill(start_color='ffF3FF00', end_color='ffF3FF00', fill_type='solid')
        for key1 in range(58, 70):
            for key2 in range(5, 7):
                wc = sheet1.cell(row=key1, column=key2)
                wc.fill = dreamteamfill

        # Creating Position symbols for GW history
        first = FormatObject(type='num', val=-1)
        second = FormatObject(type='num', val=0)
        third = FormatObject(type='num', val=1)
        iconset = IconSet(iconSet='3Arrows', cfvo=[first, second, third], showValue=None, percent=None,
                          reverse=None)
        rule = Rule(type='iconSet', iconSet=iconset)
        sheet1.conditional_formatting.add('K2:K39', cfRule=rule)

        # Creating Tables
        table1 = Table(displayName='GWH', ref='B1:L39')
        style1 = TableStyleInfo(name="TableStyleMedium11", showRowStripes=True)
        table2 = Table(displayName='GWT', ref='B41:AM56')
        style2 = TableStyleInfo(name="TableStyleLight15", showRowStripes=True)
        table1.tableStyleInfo = style1
        table2.tableStyleInfo = style2
        sheet1.add_table(table1)
        sheet1.add_table(table2)

        # Creating Charts
        c1 = LineChart()
        c1.title = 'Overall Points / Average Points / Points Benched'
        data1 = Reference(sheet1, min_col=3, max_col=5, min_row=1, max_row=39)
        c1.height = 18
        c1.width = 38
        c1.add_data(data1, titles_from_data=True)
        sheet1.add_chart(c1, "O2")

        # Save workbook
        wb.save('FPL.Data.17-18.xlsx')

        # Clear text box and show success dialog
        self.clear()
        messagebox.showinfo(title='Success', message='Check directory for Excel Workbook')

    # Define function to clear dialog
    def clear(self):
        self.fpl_prompt.delete(0, 'end')


# Call GUI window
def main():
    root = Tk()
    fantasypl(root)
    root.mainloop()


if __name__ == "__main__":
    main()
